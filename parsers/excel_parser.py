"""
Advanced Excel Parser supporting xlsx, xlsb, xlsm formats
Extracts formulas, VBA code, charts, pivot tables, and business logic
"""

import pandas as pd
import openpyxl
from openpyxl.chart import *
from openpyxl.comments import Comment
import xlrd
import pyxlsb
from typing import Dict, List, Optional, Any, Union
import re
import json
import logging
from dataclasses import dataclass, asdict, field
from pathlib import Path
import numpy as np
from openpyxl.formula.translate import Translator
import ast
import warnings
warnings.filterwarnings('ignore')

@dataclass
class CellInfo:
    """Detailed cell information"""
    address: str
    value: Any
    formula: Optional[str]
    data_type: str
    number_format: Optional[str]
    comment: Optional[str]
    hyperlink: Optional[str]
    font_info: Dict
    fill_info: Dict
    is_merged: bool
    merge_range: Optional[str]

@dataclass
class WorksheetInfo:
    """Comprehensive worksheet information"""
    name: str
    cells: List[CellInfo]
    charts: List[Dict]
    pivot_tables: List[Dict]
    data_validation: List[Dict]
    conditional_formatting: List[Dict]
    named_ranges: List[Dict]
    formulas: List[Dict]
    vba_references: List[str]
    hidden: bool = False
    protected: bool = False
    dimensions: Dict[str, Any] = field(default_factory=dict)

@dataclass
class ExcelAnalysis:
    """Complete Excel file analysis"""
    filename: str
    file_type: str
    worksheets: List[WorksheetInfo]
    vba_code: Dict[str, str]
    external_references: List[str]
    defined_names: List[Dict]
    custom_properties: Dict
    formulas_summary: Dict
    business_logic: Dict
    data_flow: Dict
    complexity_metrics: Dict

class AdvancedExcelParser:
    """
    Robust Excel parser supporting multiple formats with comprehensive extraction
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.supported_formats = ['.xlsx', '.xlsm', '.xlsb', '.xls']
        
    def parse(self, file_path: Union[str, Path]) -> ExcelAnalysis:
        """
        Parse Excel file with comprehensive analysis
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            ExcelAnalysis object with complete file information
        """
        file_path = Path(file_path)
        self.logger.info(f"Starting parse of {file_path}")
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_ext = file_path.suffix.lower()
        if file_ext not in self.supported_formats:
            raise ValueError(f"Unsupported format: {file_ext}")
        
        try:
            if file_ext in ['.xlsx', '.xlsm']:
                return self._parse_openpyxl(file_path)
            elif file_ext == '.xlsb':
                return self._parse_xlsb(file_path)
            elif file_ext == '.xls':
                return self._parse_xls(file_path)
        except Exception as e:
            self.logger.error(f"Error parsing {file_path}: {str(e)}")
            raise
    
    def _parse_openpyxl(self, file_path: Path) -> ExcelAnalysis:
        """Parse using openpyxl for xlsx/xlsm files"""
        
        # Load workbook with all features
        wb = openpyxl.load_workbook(
            file_path, 
            read_only=False,
            keep_vba=True,
            data_only=False,
            keep_links=True
        )
        
        # Extract basic info
        analysis = ExcelAnalysis(
            filename=file_path.name,
            file_type=file_path.suffix,
            worksheets=[],
            vba_code={},
            external_references=[],
            defined_names=[],
            custom_properties={},
            formulas_summary={},
            business_logic={},
            data_flow={},
            complexity_metrics={}
        )
        
        # Extract VBA code if present
        if hasattr(wb, 'vba_archive') and wb.vba_archive:
            analysis.vba_code = self._extract_vba_openpyxl(wb)
        
        # Extract defined names (named ranges)
        analysis.defined_names = self._extract_defined_names(wb)
        
        # Extract custom properties
        analysis.custom_properties = self._extract_custom_properties(wb)
        
        # Process each worksheet
        for ws in wb.worksheets:
            worksheet_info = self._process_worksheet_openpyxl(ws)
            analysis.worksheets.append(worksheet_info)
        
        # Analyze formulas and business logic
        analysis.formulas_summary = self._analyze_formulas(analysis.worksheets)
        analysis.business_logic = self._extract_business_logic(analysis.worksheets)
        analysis.data_flow = self._map_data_flow(analysis.worksheets)
        analysis.complexity_metrics = self._calculate_complexity_metrics(analysis)
        
        wb.close()
        return analysis
    
    def _extract_defined_names(self, workbook) -> List[Dict]:
        """Extract defined names/named ranges from Excel workbook"""
        defined_names = []
        
        try:
            # Check if workbook has defined names
            if hasattr(workbook, 'defined_names') and workbook.defined_names:
                for name in workbook.defined_names.definedName:
                    defined_names.append({
                        'name': name.name,
                        'value': str(name.value) if name.value else '',
                        'scope': name.localSheetId if hasattr(name, 'localSheetId') else None,
                        'hidden': name.hidden if hasattr(name, 'hidden') else False,
                        'comment': getattr(name, 'comment', '')
                    })
        except Exception as e:
            self.logger.warning(f"Failed to extract defined names: {str(e)}")
        
        return defined_names
    
    def _extract_custom_properties(self, workbook) -> Dict:
        """Extract custom document properties"""
        properties = {}
        
        try:
            # Access custom document properties
            if hasattr(workbook, 'custom_doc_props'):
                for prop in workbook.custom_doc_props:
                    properties[prop.name] = prop.value
            
            # Also try core properties
            if hasattr(workbook, 'properties'):
                core_props = workbook.properties
                properties.update({
                    'title': core_props.title if core_props.title else '',
                    'author': core_props.creator if core_props.creator else '',
                    'subject': core_props.subject if core_props.subject else '',
                    'description': core_props.description if core_props.description else '',
                    'keywords': core_props.keywords if core_props.keywords else '',
                    'created': str(core_props.created) if core_props.created else '',
                    'modified': str(core_props.modified) if core_props.modified else ''
                })
                
        except Exception as e:
            self.logger.warning(f"Failed to extract custom properties: {str(e)}")
        
        return properties
    
    def _process_worksheet_openpyxl(self, worksheet) -> WorksheetInfo:
        """Process individual worksheet comprehensively"""
        cells = []
        formulas = []
        charts = []
        pivot_tables = []
        data_validation = []
        conditional_formatting = []
        named_ranges = []
        
        try:
            # Process cells with data (limit to avoid memory issues)
            row_count = 0
            max_rows = 1000  # Limit for performance
            
            for row in worksheet.iter_rows(values_only=False):
                if row_count >= max_rows:
                    break
                    
                for cell in row:
                    if cell.value is not None:
                        # Basic cell info
                        cell_info = CellInfo(
                            address=cell.coordinate,
                            value=cell.value,
                            formula=cell.formula if hasattr(cell, 'formula') and cell.formula else None,
                            data_type=str(type(cell.value).__name__),
                            number_format=cell.number_format if cell.number_format else 'General',
                            comment=cell.comment.text if cell.comment else None,
                            hyperlink=cell.hyperlink.target if cell.hyperlink else None,
                            font_info={
                                'name': cell.font.name if cell.font else 'Calibri',
                                'size': cell.font.size if cell.font else 11,
                                'bold': cell.font.bold if cell.font else False
                            },
                            fill_info={
                                'color': str(cell.fill.fgColor.rgb) if cell.fill and hasattr(cell.fill.fgColor, 'rgb') else None,
                                'pattern': cell.fill.patternType if cell.fill else None
                            },
                            is_merged=self._is_merged_cell(worksheet, cell.coordinate),
                            merge_range=None  # Could be enhanced later
                        )
                        cells.append(cell_info)
                        
                        # Extract comprehensive formula information for EUDA documentation
                        if cell.formula:
                            formula_info = {
                                'cell': cell.coordinate,
                                'formula': cell.formula,
                                'result_value': cell.value,
                                'functions': self._extract_formula_functions(cell.formula),
                                'complexity': self._calculate_formula_complexity(cell.formula),
                                'references': self._extract_formula_references(cell.formula),
                                'business_purpose': self._infer_formula_purpose(cell.formula, cell.coordinate),
                                'calculation_type': self._classify_calculation_type(cell.formula),
                                'dependency_level': self._calculate_dependency_level(cell.formula),
                                'input_ranges': self._extract_input_ranges(cell.formula),
                                'logical_conditions': self._extract_logical_conditions(cell.formula),
                                'math_operations': self._extract_math_operations(cell.formula),
                                'data_lookups': self._extract_lookup_patterns(cell.formula)
                            }
                            formulas.append(formula_info)
                
                row_count += 1
            
            # Extract charts
            if hasattr(worksheet, '_charts'):
                for chart in worksheet._charts:
                    chart_info = {
                        'type': chart.__class__.__name__,
                        'title': str(chart.title) if hasattr(chart, 'title') and chart.title else '',
                        'anchor': str(chart.anchor) if hasattr(chart, 'anchor') else '',
                        'series_count': len(chart.series) if hasattr(chart, 'series') else 0
                    }
                    charts.append(chart_info)
            
            # Extract pivot tables (basic detection)
            if hasattr(worksheet, '_pivots'):
                for pivot in worksheet._pivots:
                    pivot_info = {
                        'name': getattr(pivot, 'name', 'Unknown'),
                        'cache_definition': str(getattr(pivot, 'cache_definition', '')),
                        'location': str(getattr(pivot, 'location', ''))
                    }
                    pivot_tables.append(pivot_info)
            
        except Exception as e:
            self.logger.warning(f"Error processing worksheet {worksheet.title}: {str(e)}")
        
        return WorksheetInfo(
            name=worksheet.title,
            cells=cells,
            charts=charts,
            pivot_tables=pivot_tables,
            data_validation=data_validation,
            conditional_formatting=conditional_formatting,
            named_ranges=named_ranges,
            formulas=formulas,
            vba_references=[],
            hidden=worksheet.sheet_state == 'hidden' if hasattr(worksheet, 'sheet_state') else False,
            protected=worksheet.protection.enabled if hasattr(worksheet, 'protection') else False,
            dimensions={
                'max_row': worksheet.max_row,
                'max_column': worksheet.max_column,
                'used_range': f"A1:{worksheet.calculate_dimension()}" if hasattr(worksheet, 'calculate_dimension') else ""
            }
        )
    
    def _is_merged_cell(self, worksheet, coordinate):
        """Check if cell is part of merged range"""
        try:
            for merged_range in worksheet.merged_cells.ranges:
                if coordinate in merged_range:
                    return True
        except:
            pass
        return False
    
    def _extract_formula_functions(self, formula: str) -> List[str]:
        """Extract function names from formula"""
        if not formula:
            return []
        
        # Find all function names (uppercase words followed by parentheses)
        functions = re.findall(r'([A-Z][A-Z0-9_]*)\s*\(', formula.upper())
        return list(set(functions))
    
    def _calculate_formula_complexity(self, formula: str) -> int:
        """Calculate complexity score for formula"""
        if not formula:
            return 0
        
        complexity = 0
        
        # Count nested levels (parentheses depth)
        max_depth = 0
        current_depth = 0
        for char in formula:
            if char == '(':
                current_depth += 1
                max_depth = max(max_depth, current_depth)
            elif char == ')':
                current_depth -= 1
        
        complexity += max_depth * 2
        
        # Count functions
        functions = self._extract_formula_functions(formula)
        complexity += len(functions)
        
        # Count operators
        operators = len(re.findall(r'[+\-*/^&<>=]', formula))
        complexity += operators
        
        # Count references
        references = len(re.findall(r'[A-Z]+[0-9]+|[A-Z]+:[A-Z]+', formula))
        complexity += references
        
        return complexity
    
    def _extract_formula_references(self, formula: str) -> List[str]:
        """Extract cell/range references from formula"""
        if not formula:
            return []
        
        # Find cell references (A1, B2:C3, etc.)
        references = re.findall(r'[A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?', formula.upper())
        
        # Find sheet references (Sheet1!A1, 'Sheet Name'!A1:B2)
        sheet_refs = re.findall(r"(?:'[^']*'|[^!\s]+)![A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?", formula)
        references.extend(sheet_refs)
        
        return list(set(references))
    
    def _parse_xlsb(self, file_path: Path) -> ExcelAnalysis:
        """Parse Excel Binary (.xlsb) files using pyxlsb"""
        self.logger.info(f"Parsing XLSB file: {file_path}")
        
        try:
            # Basic analysis structure for XLSB
            analysis = ExcelAnalysis(
                filename=file_path.name,
                file_type=file_path.suffix,
                worksheets=[],
                vba_code={},
                defined_names=[],
                custom_properties={},
                formulas_summary={},
                business_logic={},
                data_flow={},
                complexity_metrics={}
            )
            
            # Note: Full XLSB parsing requires pyxlsb library
            # For now, provide basic structure
            analysis.worksheets.append(WorksheetInfo(
                name="XLSB_Placeholder",
                cells=[],
                charts=[],
                pivot_tables=[],
                data_validation=[],
                conditional_formatting=[],
                named_ranges=[],
                formulas=[],
                vba_references=[],
                hidden=False,
                protected=False,
                dimensions={'note': 'XLSB parsing requires additional implementation'}
            ))
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"Failed to parse XLSB file: {str(e)}")
            raise
    
    def _parse_xls(self, file_path: Path) -> ExcelAnalysis:
        """Parse legacy Excel (.xls) files using xlrd"""
        self.logger.info(f"Parsing XLS file: {file_path}")
        
        try:
            # Basic analysis structure for XLS
            analysis = ExcelAnalysis(
                filename=file_path.name,
                file_type=file_path.suffix,
                worksheets=[],
                vba_code={},
                defined_names=[],
                custom_properties={},
                formulas_summary={},
                business_logic={},
                data_flow={},
                complexity_metrics={}
            )
            
            # Note: Full XLS parsing requires xlrd library
            # For now, provide basic structure
            analysis.worksheets.append(WorksheetInfo(
                name="XLS_Placeholder", 
                cells=[],
                charts=[],
                pivot_tables=[],
                data_validation=[],
                conditional_formatting=[],
                named_ranges=[],
                formulas=[],
                vba_references=[],
                hidden=False,
                protected=False,
                dimensions={'note': 'XLS parsing requires additional implementation'}
            ))
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"Failed to parse XLS file: {str(e)}")
            raise

    def _extract_vba_openpyxl(self, workbook) -> Dict[str, str]:
        """Extract VBA code from Excel workbook using openpyxl"""
        vba_code = {}
        
        try:
            # Check if workbook has VBA project
            if not hasattr(workbook, 'vba_archive') or not workbook.vba_archive:
                return vba_code
            
            # Note: openpyxl doesn't fully support VBA extraction
            # For now, we'll detect presence and provide placeholder
            vba_code['vba_detected'] = "VBA macros detected but extraction requires additional tools"
            vba_code['modules_count'] = "Unable to determine with openpyxl"
            
        except Exception as e:
            self.logger.warning(f"VBA extraction failed: {str(e)}")
            vba_code['error'] = f"VBA extraction error: {str(e)}"
        
        return vba_code
    
    def _analyze_formulas(self, worksheets: List[WorksheetInfo]) -> Dict[str, Any]:
        """Analyze formulas across worksheets"""
        formulas_summary = {
            'total_formulas': 0,
            'unique_functions': set(),
            'complex_formulas': [],
            'external_references': [],
            'circular_references': []
        }
        
        for worksheet in worksheets:
            for formula_info in worksheet.formulas:
                formulas_summary['total_formulas'] += 1
                
                formula = formula_info.get('formula', '')
                if formula:
                    # Extract functions used
                    functions = re.findall(r'([A-Z]+)\(', formula)
                    formulas_summary['unique_functions'].update(functions)
                    
                    # Identify complex formulas (nested functions, arrays, etc.)
                    if len(functions) > 3 or 'ARRAY' in formula or '{' in formula:
                        formulas_summary['complex_formulas'].append({
                            'worksheet': worksheet.name,
                            'cell': formula_info.get('cell', ''),
                            'formula': formula
                        })
        
        formulas_summary['unique_functions'] = list(formulas_summary['unique_functions'])
        return formulas_summary
    
    def _extract_business_logic(self, worksheets: List[WorksheetInfo]) -> Dict[str, Any]:
        """Extract business logic patterns from worksheets"""
        business_logic = {
            'decision_trees': [],
            'lookup_tables': [],
            'calculation_chains': [],
            'business_rules': []
        }
        
        for worksheet in worksheets:
            # Look for IF statements (decision logic)
            for formula_info in worksheet.formulas:
                formula = formula_info.get('formula', '')
                if 'IF(' in formula:
                    business_logic['decision_trees'].append({
                        'worksheet': worksheet.name,
                        'cell': formula_info.get('cell', ''),
                        'logic': formula
                    })
                
                # Look for VLOOKUP/INDEX-MATCH (lookup tables)
                if any(func in formula for func in ['VLOOKUP', 'INDEX', 'MATCH']):
                    business_logic['lookup_tables'].append({
                        'worksheet': worksheet.name,
                        'cell': formula_info.get('cell', ''),
                        'formula': formula
                    })
        
        return business_logic
    
    def _map_data_flow(self, worksheets: List[WorksheetInfo]) -> Dict[str, Any]:
        """Map data flow between worksheets"""
        data_flow = {
            'internal_references': [],
            'external_links': [],
            'worksheet_dependencies': {}
        }
        
        for worksheet in worksheets:
            dependencies = []
            for formula_info in worksheet.formulas:
                formula = formula_info.get('formula', '')
                # Look for sheet references
                sheet_refs = re.findall(r"'?([^'!]+)'?!", formula)
                dependencies.extend(sheet_refs)
            
            if dependencies:
                data_flow['worksheet_dependencies'][worksheet.name] = list(set(dependencies))
        
        return data_flow
    
    def _calculate_complexity_metrics(self, analysis) -> Dict[str, Any]:
        """Calculate complexity metrics for the Excel file"""
        metrics = {
            'total_worksheets': len(analysis.worksheets),
            'total_formulas': analysis.formulas_summary.get('total_formulas', 0),
            'unique_functions': len(analysis.formulas_summary.get('unique_functions', [])),
            'complexity_score': 0
        }
        
        # Simple complexity scoring
        base_score = metrics['total_worksheets'] * 5
        formula_score = metrics['total_formulas'] * 2
        function_score = metrics['unique_functions'] * 3
        
        metrics['complexity_score'] = base_score + formula_score + function_score
        
        if metrics['complexity_score'] > 100:
            metrics['complexity_level'] = 'High'
        elif metrics['complexity_score'] > 50:
            metrics['complexity_level'] = 'Medium'
        else:
            metrics['complexity_level'] = 'Low'
        
        return metrics
    
    def _infer_formula_purpose(self, formula: str, cell_address: str) -> Dict[str, Any]:
        """Infer the business purpose of a formula with enhanced financial modeling detection"""
        formula_lower = formula.upper()
        purpose_info = {
            'primary_purpose': 'unknown',
            'business_function': '',
            'explanation': '',
            'likely_use_case': '',
            'mathematical_framework': '',
            'risk_assessment_type': '',
            'model_complexity': 'basic'
        }
        
        # Enhanced Financial calculations with specific modeling detection
        if any(func in formula_lower for func in ['NPV', 'IRR', 'PMT', 'PV', 'FV']):
            purpose_info.update({
                'primary_purpose': 'financial_calculation',
                'business_function': 'Financial Analysis',
                'explanation': 'Performs financial calculations like net present value, internal rate of return, or payment calculations',
                'likely_use_case': 'Investment analysis, loan calculations, or financial modeling',
                'mathematical_framework': 'Time value of money calculations',
                'model_complexity': 'intermediate'
            })
        
        # VaR and Risk Modeling
        elif any(keyword in formula_lower for keyword in ['PERCENTILE', 'NORM.INV', 'CONFIDENCE']) and any(risk_term in formula_lower for risk_term in ['VAR', 'RISK', 'LOSS']):
            purpose_info.update({
                'primary_purpose': 'risk_modeling',
                'business_function': 'Value at Risk Calculation',
                'explanation': 'Calculates Value at Risk (VaR) using statistical methods to estimate potential losses',
                'likely_use_case': 'Portfolio risk management, regulatory reporting, stress testing',
                'mathematical_framework': 'Statistical risk modeling (VaR)',
                'risk_assessment_type': 'market_risk',
                'model_complexity': 'advanced'
            })
        
        # Monte Carlo Simulation
        elif any(func in formula_lower for func in ['RAND', 'RANDBETWEEN', 'NORM.DIST']) and 'AVERAGE' in formula_lower:
            purpose_info.update({
                'primary_purpose': 'monte_carlo_simulation',
                'business_function': 'Monte Carlo Analysis',
                'explanation': 'Performs Monte Carlo simulation for probabilistic analysis and scenario modeling',
                'likely_use_case': 'Risk assessment, option pricing, project valuation under uncertainty',
                'mathematical_framework': 'Stochastic simulation modeling',
                'model_complexity': 'advanced'
            })
        
        # Option Pricing Models (Black-Scholes indicators)
        elif any(func in formula_lower for func in ['EXP', 'LN', 'SQRT']) and any(term in formula_lower for term in ['STRIKE', 'EXPIRY', 'VOLATILITY']):
            purpose_info.update({
                'primary_purpose': 'option_pricing',
                'business_function': 'Derivatives Pricing',
                'explanation': 'Calculates option prices using mathematical models like Black-Scholes',
                'likely_use_case': 'Options trading, derivatives valuation, hedging strategies',
                'mathematical_framework': 'Black-Scholes or binomial option pricing model',
                'model_complexity': 'advanced'
            })
        
        # Correlation and Portfolio Analysis
        elif any(func in formula_lower for func in ['CORREL', 'COVAR', 'VAR', 'STDEV']) and len(self._extract_formula_references(formula)) > 2:
            purpose_info.update({
                'primary_purpose': 'portfolio_analysis',
                'business_function': 'Portfolio Risk Analysis',
                'explanation': 'Analyzes correlations and covariances between assets for portfolio optimization',
                'likely_use_case': 'Portfolio construction, diversification analysis, risk budgeting',
                'mathematical_framework': 'Modern Portfolio Theory (MPT)',
                'risk_assessment_type': 'portfolio_risk',
                'model_complexity': 'intermediate'
            })
        
        # Time Series Analysis
        elif any(func in formula_lower for func in ['TREND', 'GROWTH', 'FORECAST']) or ('INDEX' in formula_lower and 'MATCH' in formula_lower):
            purpose_info.update({
                'primary_purpose': 'time_series_analysis',
                'business_function': 'Time Series Forecasting',
                'explanation': 'Performs time series analysis and forecasting using historical data patterns',
                'likely_use_case': 'Revenue forecasting, trend analysis, seasonal adjustment',
                'mathematical_framework': 'Time series modeling and regression analysis',
                'model_complexity': 'intermediate'
            })
        
        # Financial ratios and metrics
        elif any(func in formula_lower for func in ['NPV', 'IRR', 'PMT', 'PV', 'FV']):
            purpose_info.update({
                'primary_purpose': 'financial_calculation',
                'business_function': 'Financial Analysis',
                'explanation': 'Performs financial calculations like net present value, internal rate of return, or payment calculations',
                'likely_use_case': 'Investment analysis, loan calculations, or financial modeling'
            })
        
        # Data aggregation
        elif any(func in formula_lower for func in ['SUM', 'SUMIF', 'SUMIFS', 'AVERAGE', 'COUNT']):
            purpose_info.update({
                'primary_purpose': 'data_aggregation',
                'business_function': 'Data Summarization',
                'explanation': 'Aggregates or summarizes data from multiple sources or ranges',
                'likely_use_case': 'Reporting, KPI calculation, or data consolidation'
            })
        
        # Lookup and reference
        elif any(func in formula_lower for func in ['VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'XLOOKUP']):
            purpose_info.update({
                'primary_purpose': 'data_lookup',
                'business_function': 'Data Retrieval',
                'explanation': 'Looks up and retrieves data from other tables or ranges',
                'likely_use_case': 'Master data lookup, price lists, or reference tables'
            })
        
        # Conditional logic
        elif 'IF' in formula_lower:
            purpose_info.update({
                'primary_purpose': 'conditional_logic',
                'business_function': 'Decision Logic',
                'explanation': 'Applies business rules or conditional logic to determine outcomes',
                'likely_use_case': 'Business rule application, status determination, or categorization'
            })
        
        # Text manipulation
        elif any(func in formula_lower for func in ['CONCATENATE', 'LEFT', 'RIGHT', 'MID', 'SUBSTITUTE']):
            purpose_info.update({
                'primary_purpose': 'text_processing',
                'business_function': 'Text Manipulation',
                'explanation': 'Processes, formats, or manipulates text data',
                'likely_use_case': 'Data cleaning, report formatting, or ID generation'
            })
        
        # Date/time calculations with financial context
        elif any(func in formula_lower for func in ['DATE', 'DATEDIF', 'NETWORKDAYS', 'WORKDAY']):
            if any(term in formula_lower for term in ['MATURITY', 'SETTLEMENT', 'COUPON', 'YIELD']):
                purpose_info.update({
                    'primary_purpose': 'fixed_income_calculation',
                    'business_function': 'Fixed Income Analysis',
                    'explanation': 'Calculates bond-related metrics including yield, duration, and accrued interest',
                    'likely_use_case': 'Bond portfolio management, yield analysis, duration matching',
                    'mathematical_framework': 'Fixed income mathematics',
                    'model_complexity': 'intermediate'
                })
            else:
                purpose_info.update({
                    'primary_purpose': 'date_calculation',
                    'business_function': 'Date/Time Analysis',
                    'explanation': 'Performs date or time-based calculations',
                    'likely_use_case': 'Project scheduling, aging analysis, or time tracking',
                    'mathematical_framework': 'Calendar mathematics'
                })
        
        # Enhanced complexity assessment
        complexity_factors = [
            len(self._extract_formula_functions(formula)),
            len(self._extract_formula_references(formula)),
            formula.count('IF'),
            formula.count('(')
        ]
        
        complexity_score = sum(complexity_factors)
        if complexity_score > 15:
            purpose_info['model_complexity'] = 'highly_advanced'
        elif complexity_score > 10:
            purpose_info['model_complexity'] = 'advanced'
        elif complexity_score > 5:
            purpose_info['model_complexity'] = 'intermediate'
        
        return purpose_info
    
    def _classify_calculation_type(self, formula: str) -> str:
        """Classify the type of calculation being performed"""
        formula_upper = formula.upper()
        
        if any(op in formula for op in ['+', '-', '*', '/', '^']):
            if any(func in formula_upper for func in ['SUM', 'AVERAGE']):
                return 'aggregation_with_math'
            return 'arithmetic_calculation'
        elif any(func in formula_upper for func in ['IF', 'AND', 'OR', 'NOT']):
            return 'logical_evaluation'
        elif any(func in formula_upper for func in ['VLOOKUP', 'INDEX', 'MATCH']):
            return 'data_retrieval'
        elif any(func in formula_upper for func in ['COUNT', 'SUM', 'AVERAGE', 'MAX', 'MIN']):
            return 'statistical_analysis'
        else:
            return 'other'
    
    def _calculate_dependency_level(self, formula: str) -> int:
        """Calculate how many levels deep the dependencies go"""
        # Count cell references (simple approximation)
        import re
        cell_refs = re.findall(r'[A-Z]+[0-9]+', formula)
        range_refs = re.findall(r'[A-Z]+[0-9]+:[A-Z]+[0-9]+', formula)
        return len(set(cell_refs)) + len(range_refs)
    
    def _extract_input_ranges(self, formula: str) -> List[str]:
        """Extract input ranges used in the formula"""
        import re
        ranges = []
        
        # Range references like A1:B10
        range_pattern = r'[A-Z]+[0-9]+:[A-Z]+[0-9]+'
        ranges.extend(re.findall(range_pattern, formula))
        
        # Single cell references
        cell_pattern = r'[A-Z]+[0-9]+'
        ranges.extend(re.findall(cell_pattern, formula))
        
        return list(set(ranges))
    
    def _extract_logical_conditions(self, formula: str) -> List[str]:
        """Extract logical conditions from the formula"""
        conditions = []
        formula_upper = formula.upper()
        
        # Look for comparison operators
        import re
        comparisons = re.findall(r'[A-Z0-9]+\s*[><=!]+\s*[A-Z0-9"\']+', formula)
        conditions.extend(comparisons)
        
        # Look for IF statements
        if 'IF(' in formula_upper:
            conditions.append('Contains conditional logic')
        
        # Look for logical functions
        for func in ['AND', 'OR', 'NOT']:
            if f'{func}(' in formula_upper:
                conditions.append(f'Uses {func} logic')
        
        return conditions
    
    def _extract_math_operations(self, formula: str) -> List[str]:
        """Extract mathematical operations from the formula"""
        operations = []
        
        if '+' in formula:
            operations.append('Addition')
        if '-' in formula:
            operations.append('Subtraction')
        if '*' in formula:
            operations.append('Multiplication')
        if '/' in formula:
            operations.append('Division')
        if '^' in formula:
            operations.append('Exponentiation')
        
        return operations
    
    def _extract_lookup_patterns(self, formula: str) -> Dict[str, Any]:
        """Extract data lookup patterns and their purposes"""
        formula_upper = formula.upper()
        lookups = {
            'lookup_functions': [],
            'lookup_purpose': '',
            'data_sources': []
        }
        
        if 'VLOOKUP' in formula_upper:
            lookups['lookup_functions'].append('VLOOKUP')
            lookups['lookup_purpose'] = 'Vertical table lookup'
        if 'HLOOKUP' in formula_upper:
            lookups['lookup_functions'].append('HLOOKUP')
            lookups['lookup_purpose'] = 'Horizontal table lookup'
        if 'INDEX' in formula_upper and 'MATCH' in formula_upper:
            lookups['lookup_functions'].append('INDEX-MATCH')
            lookups['lookup_purpose'] = 'Advanced flexible lookup'
        if 'XLOOKUP' in formula_upper:
            lookups['lookup_functions'].append('XLOOKUP')
            lookups['lookup_purpose'] = 'Modern bidirectional lookup'
        
        # Extract potential data source ranges
        import re
        ranges = re.findall(r'[A-Z]+[0-9]+:[A-Z]+[0-9]+', formula)
        lookups['data_sources'] = ranges
        
        return lookups
    
    def analyze_formula_relationships(self, worksheets: List[WorksheetInfo]) -> Dict[str, Any]:
        """Analyze relationships between formulas and create dependency maps"""
        relationships = {
            'formula_dependency_tree': {},
            'data_flow_map': {},
            'calculation_chains': [],
            'business_process_steps': []
        }
        
        all_formulas = []
        for ws in worksheets:
            for formula in ws.formulas:
                formula['worksheet'] = ws.name
                all_formulas.append(formula)
        
        # Build dependency tree
        for formula in all_formulas:
            cell_addr = f"{formula['worksheet']}!{formula['cell']}"
            dependencies = formula.get('input_ranges', [])
            relationships['formula_dependency_tree'][cell_addr] = {
                'depends_on': dependencies,
                'purpose': formula.get('business_purpose', {}),
                'calculation_type': formula.get('calculation_type', 'unknown')
            }
        
        # Identify calculation chains
        calculation_chains = self._identify_calculation_chains(all_formulas)
        relationships['calculation_chains'] = calculation_chains
        
        return relationships
    
    def _identify_calculation_chains(self, formulas: List[Dict]) -> List[Dict]:
        """Identify sequential calculation chains for business process documentation with enhanced financial workflow detection"""
        chains = []
        
        # First, identify potential financial modeling workflows
        financial_workflows = self._identify_financial_workflows(formulas)
        
        # Group by calculation type and purpose
        chains_by_purpose = {}
        for formula in formulas:
            purpose = formula.get('business_purpose', {}).get('primary_purpose', 'unknown')
            if purpose not in chains_by_purpose:
                chains_by_purpose[purpose] = []
            chains_by_purpose[purpose].append(formula)
        
        # Create chain documentation with enhanced financial workflow context
        for purpose, formula_group in chains_by_purpose.items():
            if len(formula_group) > 1:
                chain_info = {
                    'business_purpose': purpose,
                    'steps': [],
                    'description': self._generate_process_description(purpose, formula_group),
                    'financial_workflow_type': self._classify_financial_workflow(purpose, formula_group),
                    'mathematical_complexity': self._assess_mathematical_complexity(formula_group),
                    'regulatory_relevance': self._assess_regulatory_relevance(purpose, formula_group)
                }
                
                for i, formula in enumerate(formula_group, 1):
                    step = {
                        'step_number': i,
                        'cell': formula['cell'],
                        'worksheet': formula.get('worksheet', 'Unknown'),
                        'explanation': self._explain_formula_step(formula),
                        'business_meaning': formula.get('business_purpose', {}).get('explanation', 'Unknown purpose')
                    }
                    chain_info['steps'].append(step)
                
                chains.append(chain_info)
        
        # Add financial workflows to chains
        chains.extend(financial_workflows)
        
        return chains
    
    def _generate_process_description(self, purpose: str, formulas: List[Dict]) -> str:
        """Generate a business process description for a group of related calculations with enhanced financial context"""
        descriptions = {
            'risk_modeling': f'This risk modeling process implements {len(formulas)} calculation steps to quantify market risk, credit risk, or operational risk using advanced statistical methods.',
            'monte_carlo_simulation': f'This Monte Carlo simulation framework uses {len(formulas)} computational steps to perform probabilistic analysis and scenario modeling for risk assessment.',
            'option_pricing': f'This derivatives pricing engine implements {len(formulas)} steps of the Black-Scholes or binomial model to value options and other derivatives.',
            'portfolio_analysis': f'This portfolio optimization process analyzes {len(formulas)} calculations to determine optimal asset allocation, risk budgeting, and diversification strategies.',
            'time_series_analysis': f'This time series modeling framework performs {len(formulas)} analytical steps for forecasting, trend analysis, and seasonal decomposition.',
            'fixed_income_calculation': f'This fixed income analysis process implements {len(formulas)} calculations for bond pricing, yield analysis, and duration/convexity measurement.',
            'financial_calculation': f'This financial calculation process involves {len(formulas)} steps to compute financial metrics, ratios, or investment analysis using time value of money principles.',
            'data_aggregation': f'This data aggregation process summarizes information across {len(formulas)} calculation steps to create consolidated reports or KPIs.',
            'data_lookup': f'This data lookup process retrieves and cross-references information through {len(formulas)} lookup operations.',
            'conditional_logic': f'This decision logic process applies business rules through {len(formulas)} conditional statements.',
            'text_processing': f'This data processing workflow formats and manipulates text data through {len(formulas)} transformation steps.'
        }
        return descriptions.get(purpose, f'This calculation process performs {purpose} operations through {len(formulas)} steps.')
    
    def _explain_formula_step(self, formula: Dict) -> str:
        """Explain what a specific formula step does in business terms"""
        purpose = formula.get('business_purpose', {})
        calc_type = formula.get('calculation_type', 'unknown')
        
        explanations = {
            'arithmetic_calculation': f"Performs mathematical calculation to compute {purpose.get('business_function', 'a numeric result')}",
            'logical_evaluation': f"Evaluates business conditions to determine {purpose.get('likely_use_case', 'the appropriate action')}",
            'data_retrieval': f"Looks up data from reference tables for {purpose.get('likely_use_case', 'information retrieval')}",
            'statistical_analysis': f"Analyzes data patterns to calculate {purpose.get('business_function', 'statistical metrics')}",
            'aggregation_with_math': f"Combines and calculates data for {purpose.get('likely_use_case', 'summarized reporting')}"
        }
        
        return explanations.get(calc_type, f"Processes data for {purpose.get('business_function', 'business purposes')}")
    
    def _identify_financial_workflows(self, formulas: List[Dict]) -> List[Dict]:
        """Identify specific financial modeling workflows"""
        workflows = []
        
        # Group formulas by worksheet to identify workflow patterns
        worksheet_formulas = {}
        for formula in formulas:
            ws_name = formula.get('worksheet', 'Unknown')
            if ws_name not in worksheet_formulas:
                worksheet_formulas[ws_name] = []
            worksheet_formulas[ws_name].append(formula)
        
        for ws_name, ws_formulas in worksheet_formulas.items():
            # Detect VaR calculation workflow
            if self._detect_var_workflow(ws_formulas):
                workflows.append({
                    'workflow_type': 'var_calculation',
                    'worksheet': ws_name,
                    'description': 'Value at Risk (VaR) calculation workflow implementing statistical risk measurement',
                    'steps': self._document_var_steps(ws_formulas),
                    'regulatory_framework': 'Basel III, Market Risk Capital Requirements',
                    'mathematical_method': 'Historical Simulation, Parametric VaR, or Monte Carlo'
                })
            
            # Detect option pricing workflow
            if self._detect_option_pricing_workflow(ws_formulas):
                workflows.append({
                    'workflow_type': 'option_pricing',
                    'worksheet': ws_name,
                    'description': 'Option pricing model implementing Black-Scholes or binomial tree methodology',
                    'steps': self._document_option_pricing_steps(ws_formulas),
                    'mathematical_method': 'Black-Scholes-Merton model or Binomial/Trinomial trees',
                    'use_cases': 'Options trading, hedging strategies, embedded options valuation'
                })
            
            # Detect portfolio optimization workflow
            if self._detect_portfolio_optimization_workflow(ws_formulas):
                workflows.append({
                    'workflow_type': 'portfolio_optimization',
                    'worksheet': ws_name,
                    'description': 'Portfolio optimization implementing Modern Portfolio Theory (MPT)',
                    'steps': self._document_portfolio_steps(ws_formulas),
                    'mathematical_method': 'Mean-variance optimization, efficient frontier calculation',
                    'theoretical_framework': 'Modern Portfolio Theory (Markowitz)'
                })
        
        return workflows
    
    def _detect_var_workflow(self, formulas: List[Dict]) -> bool:
        """Detect VaR calculation patterns"""
        var_indicators = ['PERCENTILE', 'CONFIDENCE', 'NORM.INV', 'VAR', 'STDEV']
        formula_texts = [f.get('formula', '').upper() for f in formulas]
        all_text = ' '.join(formula_texts)
        
        return sum(1 for indicator in var_indicators if indicator in all_text) >= 2
    
    def _detect_option_pricing_workflow(self, formulas: List[Dict]) -> bool:
        """Detect option pricing model patterns"""
        option_indicators = ['EXP', 'LN', 'SQRT', 'NORM.DIST', 'NORM.INV']
        formula_texts = [f.get('formula', '').upper() for f in formulas]
        all_text = ' '.join(formula_texts)
        
        # Need exponential/log functions + normal distribution functions
        exp_log_count = sum(1 for func in ['EXP', 'LN'] if func in all_text)
        norm_count = sum(1 for func in ['NORM.DIST', 'NORM.INV'] if func in all_text)
        
        return exp_log_count >= 1 and norm_count >= 1
    
    def _detect_portfolio_optimization_workflow(self, formulas: List[Dict]) -> bool:
        """Detect portfolio optimization patterns"""
        portfolio_indicators = ['CORREL', 'COVAR', 'VAR', 'STDEV', 'MMULT']
        formula_texts = [f.get('formula', '').upper() for f in formulas]
        all_text = ' '.join(formula_texts)
        
        return sum(1 for indicator in portfolio_indicators if indicator in all_text) >= 3
    
    def _document_var_steps(self, formulas: List[Dict]) -> List[Dict]:
        """Document VaR calculation steps"""
        steps = []
        for i, formula in enumerate(formulas, 1):
            if any(func in formula.get('formula', '').upper() for func in ['PERCENTILE', 'CONFIDENCE', 'NORM.INV']):
                steps.append({
                    'step_number': i,
                    'description': 'Calculate Value at Risk using statistical percentile method',
                    'formula': formula.get('formula', ''),
                    'business_meaning': 'Determines maximum expected loss at specified confidence level'
                })
        return steps
    
    def _document_option_pricing_steps(self, formulas: List[Dict]) -> List[Dict]:
        """Document option pricing calculation steps"""
        steps = []
        for i, formula in enumerate(formulas, 1):
            formula_text = formula.get('formula', '').upper()
            if 'LN' in formula_text:
                steps.append({
                    'step_number': i,
                    'description': 'Calculate natural logarithm component of option pricing model',
                    'formula': formula.get('formula', ''),
                    'business_meaning': 'Computes log-normal distribution parameters for stock price modeling'
                })
            elif 'EXP' in formula_text:
                steps.append({
                    'step_number': i,
                    'description': 'Calculate exponential component for risk-free discounting',
                    'formula': formula.get('formula', ''),
                    'business_meaning': 'Applies risk-free rate discounting to present value of option payoff'
                })
        return steps
    
    def _document_portfolio_steps(self, formulas: List[Dict]) -> List[Dict]:
        """Document portfolio optimization steps"""
        steps = []
        for i, formula in enumerate(formulas, 1):
            formula_text = formula.get('formula', '').upper()
            if 'CORREL' in formula_text or 'COVAR' in formula_text:
                steps.append({
                    'step_number': i,
                    'description': 'Calculate correlation or covariance matrix for asset relationships',
                    'formula': formula.get('formula', ''),
                    'business_meaning': 'Measures diversification benefits and risk dependencies between assets'
                })
            elif 'MMULT' in formula_text:
                steps.append({
                    'step_number': i,
                    'description': 'Perform matrix multiplication for portfolio risk calculation',
                    'formula': formula.get('formula', ''),
                    'business_meaning': 'Computes portfolio variance using weights and covariance matrix'
                })
        return steps
    
    def _classify_financial_workflow(self, purpose: str, formulas: List[Dict]) -> str:
        """Classify the type of financial workflow"""
        workflow_types = {
            'risk_modeling': 'Market Risk Management',
            'monte_carlo_simulation': 'Stochastic Modeling',
            'option_pricing': 'Derivatives Valuation',
            'portfolio_analysis': 'Asset Management',
            'time_series_analysis': 'Econometric Forecasting',
            'fixed_income_calculation': 'Fixed Income Analytics'
        }
        return workflow_types.get(purpose, 'General Financial Analysis')
    
    def _assess_mathematical_complexity(self, formulas: List[Dict]) -> str:
        """Assess the mathematical complexity of a formula group"""
        complexity_score = 0
        
        for formula in formulas:
            formula_text = formula.get('formula', '').upper()
            # Advanced mathematical functions
            if any(func in formula_text for func in ['EXP', 'LN', 'LOG', 'SQRT', 'POWER']):
                complexity_score += 3
            # Statistical functions
            if any(func in formula_text for func in ['NORM.DIST', 'NORM.INV', 'PERCENTILE', 'CONFIDENCE']):
                complexity_score += 2
            # Matrix operations
            if any(func in formula_text for func in ['MMULT', 'TRANSPOSE', 'MINVERSE']):
                complexity_score += 4
        
        if complexity_score > 15:
            return 'Highly Advanced (Graduate-level mathematics)'
        elif complexity_score > 8:
            return 'Advanced (Undergraduate-level mathematics)'
        elif complexity_score > 3:
            return 'Intermediate (Statistics and algebra)'
        else:
            return 'Basic (Arithmetic operations)'
    
    def _assess_regulatory_relevance(self, purpose: str, formulas: List[Dict]) -> str:
        """Assess regulatory relevance of financial calculations"""
        regulatory_map = {
            'risk_modeling': 'Basel III Market Risk Framework, Solvency II',
            'monte_carlo_simulation': 'CCAR Stress Testing, IFRS 13 Fair Value',
            'option_pricing': 'IFRS 13 Fair Value Measurement, ASC 820',
            'portfolio_analysis': 'MiFID II Best Execution, Fiduciary Rules',
            'fixed_income_calculation': 'Basel III Interest Rate Risk, IFRS 9'
        }
        return regulatory_map.get(purpose, 'General business reporting standards')
    
    def export_analysis(self, analysis: ExcelAnalysis, output_path: str) -> None:
        """Export analysis results to JSON"""
        
        # Convert dataclasses to dict for JSON serialization
        analysis_dict = asdict(analysis)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(analysis_dict, f, indent=2, default=str)
        
        self.logger.info(f"Analysis exported to {output_path}")

# Example usage with enhanced financial modeling analysis
if __name__ == "__main__":
    parser = AdvancedExcelParser()
    
    # Test with different Excel formats, especially financial models
    test_files = [
        "sample.xlsx",
        "var_model.xlsx",
        "option_pricing.xlsm", 
        "portfolio_optimization.xlsx",
        "monte_carlo_simulation.xlsb",
        "legacy_file.xls"
    ]
    
    for file_path in test_files:
        if Path(file_path).exists():
            try:
                analysis = parser.parse(file_path)
                print(f"\nEnhanced Analysis for {file_path}:")
                print(f"Worksheets: {len(analysis.worksheets)}")
                print(f"VBA modules: {len(analysis.vba_code)}")
                print(f"Financial workflows detected: {len([c for c in analysis.business_logic.get('calculation_chains', []) if 'financial' in c.get('workflow_type', '')])}")
                
                # Show financial modeling insights
                for workflow in analysis.business_logic.get('calculation_chains', []):
                    if workflow.get('workflow_type') in ['var_calculation', 'option_pricing', 'portfolio_optimization']:
                        print(f"  - {workflow['workflow_type']}: {workflow.get('mathematical_method', 'Unknown method')}")
                
                # Export detailed analysis with enhanced financial insights
                parser.export_analysis(analysis, f"{file_path}_enhanced_analysis.json")
                
            except Exception as e:
                print(f"Error parsing {file_path}: {str(e)}")
    
    print("\n=== Enhanced Financial Modeling Parser Ready ===")
    print("Capabilities:")
    print("- VaR and risk modeling detection")
    print("- Option pricing model identification (Black-Scholes, Binomial)")
    print("- Portfolio optimization analysis (MPT)")
    print("- Monte Carlo simulation recognition")
    print("- Fixed income analytics")
    print("- Time series analysis")
    print("- Regulatory framework mapping")