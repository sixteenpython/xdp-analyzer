"""
Advanced Excel Parser
Formula-aware parsing with financial modeling recognition and VBA analysis
"""

import pandas as pd
import numpy as np
import logging
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass, field
from pathlib import Path
import re
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

try:
    import openpyxl
    from openpyxl.formula import Tokenizer
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

@dataclass
class CellInfo:
    """Information about an Excel cell"""
    coordinate: str
    value: Any
    formula: Optional[str] = None
    data_type: str = 'unknown'
    number_format: Optional[str] = None
    has_formula: bool = False
    references: List[str] = field(default_factory=list)
    precedents: List[str] = field(default_factory=list)
    dependents: List[str] = field(default_factory=list)

@dataclass
class WorksheetAnalysis:
    """Advanced analysis of an Excel worksheet"""
    name: str
    cells: List[CellInfo]
    formulas: List[str]
    charts: List[Dict[str, Any]]
    pivot_tables: List[Dict[str, Any]]
    data_validation: List[Dict[str, Any]]
    conditional_formatting: List[Dict[str, Any]]
    named_ranges: List[Dict[str, Any]]
    vba_references: List[str]
    hidden: bool
    protected: bool
    dimensions: Dict[str, int]
    # Financial modeling specific
    financial_formulas: Dict[str, List[str]] = field(default_factory=dict)
    risk_calculations: List[str] = field(default_factory=list)
    time_series_formulas: List[str] = field(default_factory=list)
    option_pricing_models: List[str] = field(default_factory=list)
    portfolio_calculations: List[str] = field(default_factory=list)

@dataclass
class VBAAnalysis:
    """VBA code analysis"""
    module_name: str
    code: str
    functions: List[str]
    subroutines: List[str]
    external_references: List[str]
    financial_functions: List[str]
    risk_management_code: List[str]
    automation_level: str  # 'none', 'basic', 'advanced'
    complexity_score: float

@dataclass
class AdvancedExcelAnalysis:
    """Complete advanced Excel analysis"""
    filename: str
    worksheets: List[WorksheetAnalysis]
    vba_code: Dict[str, VBAAnalysis]
    external_references: List[str]
    defined_names: List[Dict[str, Any]]
    custom_properties: Dict[str, Any]
    formulas_summary: Dict[str, Any]
    business_logic: Dict[str, Any]
    data_flow: Dict[str, Any]
    complexity_metrics: Dict[str, Any]
    financial_modeling_analysis: Dict[str, Any] = field(default_factory=dict)
    risk_management_assessment: Dict[str, Any] = field(default_factory=dict)
    model_validation_indicators: Dict[str, Any] = field(default_factory=dict)

class AdvancedExcelParser:
    """
    Advanced Excel parser with formula analysis and financial modeling recognition
    Requires openpyxl for formula parsing and VBA analysis
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        if not HAS_OPENPYXL:
            self.logger.warning("openpyxl not available. Advanced parsing will be limited.")
        
        # Financial modeling function patterns
        self.financial_functions = {
            'var_risk': ['VAR', 'COVAR', 'CORREL', 'STDEV', 'PERCENTILE', 'NORMDIST', 'NORMSINV'],
            'option_pricing': ['EXP', 'LN', 'SQRT', 'NORMSDIST', 'NORMSINV', 'MAX'],
            'fixed_income': ['PV', 'FV', 'PMT', 'RATE', 'NPER', 'DURATION', 'MDURATION', 'YIELD'],
            'time_series': ['TREND', 'FORECAST', 'LINEST', 'LOGEST', 'GROWTH'],
            'portfolio': ['SUMPRODUCT', 'MMULT', 'TRANSPOSE', 'MINVERSE'],
            'monte_carlo': ['RAND', 'RANDBETWEEN', 'RANDARRAY', 'NORM.INV', 'UNIFORM']
        }
        
        # Business logic patterns in formulas
        self.business_patterns = {
            'risk_management': ['IF.*RISK', 'STRESS.*TEST', 'SCENARIO.*ANALYSIS'],
            'backtesting': ['BACK.*TEST', 'HISTORICAL.*SIMULATION', 'WALK.*FORWARD'],
            'model_validation': ['VALIDATE.*', 'CHECK.*', 'VERIFY.*', 'TEST.*'],
            'sensitivity': ['SENSITIVITY.*', 'WHAT.*IF', 'SCENARIO.*'],
            'optimization': ['SOLVER', 'OPTIMIZE', 'MINIMIZE', 'MAXIMIZE']
        }
        
        # VBA financial function patterns
        self.vba_financial_patterns = {
            'advanced_modeling': ['BlackScholes', 'BinomialOption', 'MonteCarloVaR', 'GARCH'],
            'risk_systems': ['VaRCalculation', 'StressTest', 'BackTest', 'RiskReport'],
            'portfolio_management': ['PortfolioOptimization', 'AssetAllocation', 'Rebalancing'],
            'derivatives': ['OptionGreeks', 'SwapPricing', 'ForwardCurve', 'VolatilitySurface']
        }
    
    def parse(self, file_path: Union[str, Path]) -> AdvancedExcelAnalysis:
        """
        Parse Excel file with advanced formula and VBA analysis
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            AdvancedExcelAnalysis object
        """
        file_path = Path(file_path)
        self.logger.info(f"Starting advanced parse of {file_path}")
        
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for advanced Excel parsing")
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)
            
            # Analyze worksheets
            worksheets = []
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                worksheet_analysis = self._analyze_worksheet(worksheet)
                worksheets.append(worksheet_analysis)
            
            # Analyze VBA code
            vba_analysis = self._analyze_vba(workbook)
            
            # Extract defined names
            defined_names = self._extract_defined_names(workbook)
            
            # Analyze external references
            external_refs = self._extract_external_references(workbook)
            
            # Extract custom properties
            custom_props = self._extract_custom_properties(workbook)
            
            # Generate formulas summary
            formulas_summary = self._generate_formulas_summary(worksheets)
            
            # Analyze business logic
            business_logic = self._analyze_business_logic(worksheets, vba_analysis)
            
            # Map data flow
            data_flow = self._analyze_data_flow(worksheets)
            
            # Calculate complexity metrics
            complexity_metrics = self._calculate_complexity_metrics(worksheets, vba_analysis, formulas_summary)
            
            # Financial modeling analysis
            financial_analysis = self._analyze_financial_modeling(worksheets, vba_analysis, formulas_summary)
            
            # Risk management assessment
            risk_assessment = self._assess_risk_management_capabilities(worksheets, vba_analysis)
            
            # Model validation indicators
            validation_indicators = self._identify_model_validation_indicators(worksheets, vba_analysis)
            
            analysis = AdvancedExcelAnalysis(
                filename=file_path.name,
                worksheets=worksheets,
                vba_code=vba_analysis,
                external_references=external_refs,
                defined_names=defined_names,
                custom_properties=custom_props,
                formulas_summary=formulas_summary,
                business_logic=business_logic,
                data_flow=data_flow,
                complexity_metrics=complexity_metrics,
                financial_modeling_analysis=financial_analysis,
                risk_management_assessment=risk_assessment,
                model_validation_indicators=validation_indicators
            )
            
            workbook.close()
            return analysis
            
        except Exception as e:
            self.logger.error(f"Error parsing {file_path}: {str(e)}")
            raise
    
    def _analyze_worksheet(self, worksheet) -> WorksheetAnalysis:
        """Analyze individual worksheet with formula parsing"""
        
        cells = []
        formulas = []
        financial_formulas = {'var_risk': [], 'option_pricing': [], 'fixed_income': [], 'time_series': [], 'portfolio': []}
        risk_calculations = []
        time_series_formulas = []
        option_pricing_models = []
        portfolio_calculations = []
        
        # Analyze each cell
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None or cell.data_type == 'f':
                    cell_info = self._analyze_cell(cell)
                    cells.append(cell_info)
                    
                    if cell_info.has_formula:
                        formulas.append(cell_info.formula)
                        
                        # Categorize financial formulas
                        formula_upper = cell_info.formula.upper()
                        for category, functions in self.financial_functions.items():
                            if any(func in formula_upper for func in functions):
                                financial_formulas[category].append(cell_info.formula)
                        
                        # Identify specific model types
                        if any(pattern in formula_upper for pattern in ['VAR', 'PERCENTILE', 'NORMINV']):
                            risk_calculations.append(cell_info.formula)
                        
                        if any(pattern in formula_upper for pattern in ['BLACK', 'SCHOLES', 'BINOMIAL', 'OPTION']):
                            option_pricing_models.append(cell_info.formula)
                        
                        if 'SUMPRODUCT' in formula_upper and any(ref in formula_upper for ref in ['WEIGHT', 'ALLOCATION']):
                            portfolio_calculations.append(cell_info.formula)
                        
                        if any(func in formula_upper for func in ['TREND', 'FORECAST', 'LINEST']):
                            time_series_formulas.append(cell_info.formula)
        
        # Analyze charts (simplified)
        charts = self._analyze_charts(worksheet)
        
        # Analyze pivot tables (simplified)
        pivot_tables = self._analyze_pivot_tables(worksheet)
        
        # Extract other worksheet features
        data_validation = self._extract_data_validation(worksheet)
        conditional_formatting = self._extract_conditional_formatting(worksheet)
        named_ranges = self._extract_named_ranges(worksheet)
        
        return WorksheetAnalysis(
            name=worksheet.title,
            cells=cells,
            formulas=formulas,
            charts=charts,
            pivot_tables=pivot_tables,
            data_validation=data_validation,
            conditional_formatting=conditional_formatting,
            named_ranges=named_ranges,
            vba_references=[],  # Will be populated during VBA analysis
            hidden=worksheet.sheet_state == 'hidden',
            protected=worksheet.protection.sheet,
            dimensions={
                'max_row': worksheet.max_row or 0,
                'max_column': worksheet.max_column or 0
            },
            financial_formulas=financial_formulas,
            risk_calculations=risk_calculations,
            time_series_formulas=time_series_formulas,
            option_pricing_models=option_pricing_models,
            portfolio_calculations=portfolio_calculations
        )
    
    def _analyze_cell(self, cell) -> CellInfo:
        """Analyze individual cell for formula and data type"""
        
        coordinate = cell.coordinate
        value = cell.value
        formula = None
        has_formula = False
        references = []
        
        # Check if cell has formula
        if cell.data_type == 'f':
            has_formula = True
            formula = str(cell.value) if cell.value else ""
            
            # Extract cell references from formula
            if formula:
                try:
                    tokenizer = Tokenizer(formula)
                    for token in tokenizer.items:
                        if token.type == token.OPERAND and ':' in token.value:
                            references.append(token.value)
                        elif token.type == token.OPERAND and re.match(r'^[A-Z]+\d+$', token.value):
                            references.append(token.value)
                except:
                    pass
        
        return CellInfo(
            coordinate=coordinate,
            value=value,
            formula=formula,
            data_type=cell.data_type,
            number_format=cell.number_format,
            has_formula=has_formula,
            references=references
        )
    
    def _analyze_vba(self, workbook) -> Dict[str, VBAAnalysis]:
        """Analyze VBA code for financial modeling patterns"""
        
        vba_analysis = {}
        
        try:
            if hasattr(workbook, 'vba_archive') and workbook.vba_archive:
                # This is a simplified VBA analysis
                # In practice, you'd need python-vba or similar for full VBA parsing
                
                vba_analysis['general'] = VBAAnalysis(
                    module_name='VBA_PROJECT',
                    code='VBA code detected but not parsed',
                    functions=[],
                    subroutines=[],
                    external_references=[],
                    financial_functions=[],
                    risk_management_code=[],
                    automation_level='detected',
                    complexity_score=50.0  # Placeholder
                )
                
        except Exception as e:
            self.logger.warning(f"VBA analysis failed: {e}")
        
        return vba_analysis
    
    def _analyze_charts(self, worksheet) -> List[Dict[str, Any]]:
        """Analyze charts in worksheet"""
        charts = []
        
        try:
            for chart in worksheet._charts:
                chart_info = {
                    'type': type(chart).__name__,
                    'title': getattr(chart, 'title', 'Untitled'),
                    'series_count': len(getattr(chart, 'series', [])),
                    'financial_relevance': self._assess_chart_financial_relevance(chart)
                }
                charts.append(chart_info)
        except:
            pass
        
        return charts
    
    def _assess_chart_financial_relevance(self, chart) -> str:
        """Assess if chart is relevant for financial modeling"""
        
        try:
            chart_title = str(getattr(chart, 'title', '')).lower()
            
            financial_keywords = ['return', 'risk', 'volatility', 'price', 'yield', 'var', 'portfolio']
            
            if any(keyword in chart_title for keyword in financial_keywords):
                return 'high'
            elif any(keyword in chart_title for keyword in ['trend', 'performance', 'analysis']):
                return 'medium'
            else:
                return 'low'
                
        except:
            return 'unknown'
    
    def _analyze_pivot_tables(self, worksheet) -> List[Dict[str, Any]]:
        """Analyze pivot tables in worksheet"""
        pivot_tables = []
        
        try:
            for pivot in worksheet._pivots:
                pivot_info = {
                    'name': getattr(pivot, 'name', 'Unnamed'),
                    'source_range': str(getattr(pivot, 'cache', {}).get('cacheSource', '')),
                    'fields_count': len(getattr(pivot, 'fields', [])),
                    'financial_relevance': 'medium'  # Most pivots in financial models are relevant
                }
                pivot_tables.append(pivot_info)
        except:
            pass
        
        return pivot_tables
    
    def _extract_data_validation(self, worksheet) -> List[Dict[str, Any]]:
        """Extract data validation rules"""
        validation_rules = []
        
        try:
            for dv in worksheet.data_validations.dataValidation:
                rule = {
                    'type': dv.type,
                    'ranges': [str(r) for r in dv.ranges],
                    'formula1': dv.formula1,
                    'formula2': dv.formula2
                }
                validation_rules.append(rule)
        except:
            pass
        
        return validation_rules
    
    def _extract_conditional_formatting(self, worksheet) -> List[Dict[str, Any]]:
        """Extract conditional formatting rules"""
        cf_rules = []
        
        try:
            for cf in worksheet.conditional_formatting:
                rule = {
                    'type': type(cf.cfRule[0]).__name__ if cf.cfRule else 'Unknown',
                    'ranges': [str(r) for r in cf.cells],
                    'priority': getattr(cf.cfRule[0], 'priority', 0) if cf.cfRule else 0
                }
                cf_rules.append(rule)
        except:
            pass
        
        return cf_rules
    
    def _extract_named_ranges(self, worksheet) -> List[Dict[str, Any]]:
        """Extract named ranges specific to worksheet"""
        named_ranges = []
        
        try:
            workbook = worksheet.parent
            for name in workbook.defined_names.definedName:
                if name.localSheetId == worksheet._id or name.localSheetId is None:
                    named_range = {
                        'name': name.name,
                        'refers_to': name.attr_text,
                        'scope': 'worksheet' if name.localSheetId else 'workbook',
                        'financial_relevance': self._assess_named_range_relevance(name.name)
                    }
                    named_ranges.append(named_range)
        except:
            pass
        
        return named_ranges
    
    def _assess_named_range_relevance(self, name: str) -> str:
        """Assess financial relevance of named range"""
        
        name_lower = name.lower()
        
        high_relevance = ['var', 'volatility', 'correlation', 'portfolio', 'risk', 'return', 'price', 'yield']
        medium_relevance = ['data', 'input', 'output', 'parameter', 'assumption']
        
        if any(keyword in name_lower for keyword in high_relevance):
            return 'high'
        elif any(keyword in name_lower for keyword in medium_relevance):
            return 'medium'
        else:
            return 'low'
    
    def _extract_defined_names(self, workbook) -> List[Dict[str, Any]]:
        """Extract all defined names from workbook"""
        defined_names = []
        
        try:
            for name in workbook.defined_names.definedName:
                defined_name = {
                    'name': name.name,
                    'refers_to': name.attr_text,
                    'scope': 'workbook' if name.localSheetId is None else f'sheet_{name.localSheetId}',
                    'financial_relevance': self._assess_named_range_relevance(name.name)
                }
                defined_names.append(defined_name)
        except:
            pass
        
        return defined_names
    
    def _extract_external_references(self, workbook) -> List[str]:
        """Extract external references"""
        external_refs = []
        
        try:
            # This is simplified - would need deeper analysis for full external ref detection
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f' and cell.value:
                            formula = str(cell.value)
                            if '[' in formula and ']' in formula:
                                # Likely external reference
                                external_refs.append(formula)
        except:
            pass
        
        return list(set(external_refs))
    
    def _extract_custom_properties(self, workbook) -> Dict[str, Any]:
        """Extract custom document properties"""
        properties = {}
        
        try:
            if hasattr(workbook, 'custom_doc_props'):
                for prop in workbook.custom_doc_props:
                    properties[prop.name] = prop.value
        except:
            pass
        
        return properties
    
    def _generate_formulas_summary(self, worksheets: List[WorksheetAnalysis]) -> Dict[str, Any]:
        """Generate summary of formulas across all worksheets"""
        
        total_formulas = sum(len(ws.formulas) for ws in worksheets)
        all_formulas = []
        for ws in worksheets:
            all_formulas.extend(ws.formulas)
        
        # Extract unique functions
        functions_used = set()
        for formula in all_formulas:
            formula_upper = formula.upper()
            # Simple function extraction
            import re
            funcs = re.findall(r'([A-Z][A-Z0-9]*)\s*\(', formula_upper)
            functions_used.update(funcs)
        
        # Categorize functions by financial modeling area
        function_categories = {}
        for category, category_funcs in self.financial_functions.items():
            category_count = sum(1 for func in functions_used if func in category_funcs)
            if category_count > 0:
                function_categories[category] = category_count
        
        # Calculate complexity indicators
        avg_formula_length = np.mean([len(f) for f in all_formulas]) if all_formulas else 0
        nested_formula_count = sum(1 for f in all_formulas if f.count('(') > 2)
        
        return {
            'total_formulas': total_formulas,
            'unique_functions': list(functions_used),
            'function_count': len(functions_used),
            'financial_function_categories': function_categories,
            'average_formula_length': avg_formula_length,
            'nested_formulas_count': nested_formula_count,
            'complexity_score': min(100, avg_formula_length * 0.5 + nested_formula_count * 2)
        }
    
    def _analyze_business_logic(self, worksheets: List[WorksheetAnalysis], vba_analysis: Dict[str, VBAAnalysis]) -> Dict[str, Any]:
        """Analyze business logic patterns"""
        
        business_logic = {
            'financial_modeling_patterns': [],
            'risk_management_logic': [],
            'data_validation_rules': [],
            'automation_level': 'manual',
            'model_sophistication': 'basic'
        }
        
        # Analyze formula patterns
        all_formulas = []
        for ws in worksheets:
            all_formulas.extend(ws.formulas)
            business_logic['data_validation_rules'].extend(ws.data_validation)
        
        # Look for business logic patterns in formulas
        for pattern_type, patterns in self.business_patterns.items():
            for formula in all_formulas:
                formula_upper = formula.upper()
                for pattern in patterns:
                    if re.search(pattern, formula_upper):
                        business_logic['financial_modeling_patterns'].append({
                            'type': pattern_type,
                            'pattern': pattern,
                            'formula': formula[:100]  # Truncate for readability
                        })
        
        # Assess sophistication based on detected patterns
        pattern_count = len(business_logic['financial_modeling_patterns'])
        if pattern_count > 10:
            business_logic['model_sophistication'] = 'advanced'
        elif pattern_count > 3:
            business_logic['model_sophistication'] = 'intermediate'
        
        # Check VBA for automation
        if vba_analysis:
            business_logic['automation_level'] = 'automated'
        
        return business_logic
    
    def _analyze_data_flow(self, worksheets: List[WorksheetAnalysis]) -> Dict[str, Any]:
        """Analyze data flow between worksheets"""
        
        data_flow = {
            'cross_sheet_references': [],
            'data_sources': [],
            'calculation_chain': [],
            'circular_references': []
        }
        
        # Analyze cross-sheet references
        for ws in worksheets:
            for cell in ws.cells:
                if cell.has_formula and cell.references:
                    for ref in cell.references:
                        if '!' in ref:  # Cross-sheet reference
                            data_flow['cross_sheet_references'].append({
                                'source_sheet': ws.name,
                                'source_cell': cell.coordinate,
                                'target_reference': ref
                            })
        
        # Identify data sources (sheets with minimal incoming references)
        sheet_incoming_refs = {ws.name: 0 for ws in worksheets}
        for ref in data_flow['cross_sheet_references']:
            target_sheet = ref['target_reference'].split('!')[0] if '!' in ref['target_reference'] else None
            if target_sheet:
                sheet_incoming_refs[target_sheet] = sheet_incoming_refs.get(target_sheet, 0) + 1
        
        data_flow['data_sources'] = [sheet for sheet, count in sheet_incoming_refs.items() if count < 2]
        
        return data_flow
    
    def _calculate_complexity_metrics(self, worksheets: List[WorksheetAnalysis], 
                                    vba_analysis: Dict[str, VBAAnalysis], 
                                    formulas_summary: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate overall complexity metrics"""
        
        # Basic metrics
        total_cells_with_formulas = sum(len([c for c in ws.cells if c.has_formula]) for ws in worksheets)
        total_worksheets = len(worksheets)
        total_charts = sum(len(ws.charts) for ws in worksheets)
        total_pivot_tables = sum(len(ws.pivot_tables) for ws in worksheets)
        
        # Formula complexity
        formula_complexity = formulas_summary.get('complexity_score', 0)
        
        # VBA complexity
        vba_complexity = sum(va.complexity_score for va in vba_analysis.values()) if vba_analysis else 0
        
        # Financial modeling complexity
        financial_categories = len(formulas_summary.get('financial_function_categories', {}))
        
        # Overall complexity score
        complexity_score = (
            total_cells_with_formulas * 0.1 +
            total_worksheets * 2 +
            total_charts * 1.5 +
            total_pivot_tables * 3 +
            formula_complexity * 0.5 +
            vba_complexity * 0.3 +
            financial_categories * 10
        )
        
        return {
            'total_cells_with_formulas': total_cells_with_formulas,
            'total_worksheets': total_worksheets,
            'total_charts': total_charts,
            'total_pivot_tables': total_pivot_tables,
            'formula_complexity_score': formula_complexity,
            'vba_complexity_score': vba_complexity,
            'financial_modeling_complexity': financial_categories,
            'complexity_score': round(complexity_score, 2),
            'complexity_level': 'high' if complexity_score > 100 else 'medium' if complexity_score > 50 else 'low'
        }
    
    def _analyze_financial_modeling(self, worksheets: List[WorksheetAnalysis], 
                                  vba_analysis: Dict[str, VBAAnalysis], 
                                  formulas_summary: Dict[str, Any]) -> Dict[str, Any]:
        """Comprehensive financial modeling analysis"""
        
        analysis = {
            'model_types_detected': [],
            'sophistication_level': 'basic',
            'financial_domains': [],
            'risk_modeling_capability': False,
            'monte_carlo_simulation': False,
            'option_pricing_models': False,
            'portfolio_optimization': False,
            'time_series_analysis': False,
            'backtesting_framework': False,
            'model_validation_present': False
        }
        
        # Analyze by worksheet
        all_risk_calculations = []
        all_option_models = []
        all_portfolio_calcs = []
        all_time_series = []
        
        for ws in worksheets:
            all_risk_calculations.extend(ws.risk_calculations)
            all_option_models.extend(ws.option_pricing_models)
            all_portfolio_calcs.extend(ws.portfolio_calculations)
            all_time_series.extend(ws.time_series_formulas)
        
        # Determine model types
        if all_risk_calculations:
            analysis['model_types_detected'].append('VaR/Risk Management')
            analysis['risk_modeling_capability'] = True
            analysis['financial_domains'].append('risk_management')
        
        if all_option_models:
            analysis['model_types_detected'].append('Option Pricing')
            analysis['option_pricing_models'] = True
            analysis['financial_domains'].append('derivatives')
        
        if all_portfolio_calcs:
            analysis['model_types_detected'].append('Portfolio Management')
            analysis['portfolio_optimization'] = True
            analysis['financial_domains'].append('portfolio_management')
        
        if all_time_series:
            analysis['model_types_detected'].append('Time Series Analysis')
            analysis['time_series_analysis'] = True
            analysis['financial_domains'].append('forecasting')
        
        # Check for Monte Carlo simulation
        monte_carlo_indicators = ['RAND', 'RANDBETWEEN', 'NORM.INV', 'UNIFORM']
        has_random_functions = any(func in formulas_summary.get('unique_functions', []) 
                                 for func in monte_carlo_indicators)
        if has_random_functions and len(all_risk_calculations) > 0:
            analysis['monte_carlo_simulation'] = True
            analysis['model_types_detected'].append('Monte Carlo Simulation')
        
        # Assess sophistication
        model_count = len(analysis['model_types_detected'])
        if model_count >= 4:
            analysis['sophistication_level'] = 'advanced'
        elif model_count >= 2:
            analysis['sophistication_level'] = 'intermediate'
        
        return analysis
    
    def _assess_risk_management_capabilities(self, worksheets: List[WorksheetAnalysis], 
                                           vba_analysis: Dict[str, VBAAnalysis]) -> Dict[str, Any]:
        """Assess risk management modeling capabilities"""
        
        assessment = {
            'var_modeling': False,
            'stress_testing': False,
            'scenario_analysis': False,
            'backtesting': False,
            'risk_reporting': False,
            'regulatory_compliance': [],
            'risk_metrics_calculated': [],
            'model_validation_tests': []
        }
        
        # Analyze formulas for risk calculations
        all_formulas = []
        for ws in worksheets:
            all_formulas.extend(ws.formulas)
            all_formulas.extend(ws.risk_calculations)
        
        formula_text = ' '.join(all_formulas).upper()
        
        # VaR modeling detection
        if any(term in formula_text for term in ['VAR', 'PERCENTILE', 'NORMINV', 'CONFIDENCE']):
            assessment['var_modeling'] = True
            assessment['risk_metrics_calculated'].append('Value at Risk')
        
        # Stress testing detection
        if any(term in formula_text for term in ['STRESS', 'SCENARIO', 'WHAT.IF']):
            assessment['stress_testing'] = True
            assessment['scenario_analysis'] = True
        
        # Backtesting detection
        if any(term in formula_text for term in ['BACKTEST', 'HISTORICAL', 'VALIDATION']):
            assessment['backtesting'] = True
            assessment['model_validation_tests'].append('Historical Backtesting')
        
        # Additional risk metrics
        if 'CORREL' in formula_text or 'COVAR' in formula_text:
            assessment['risk_metrics_calculated'].append('Correlation Analysis')
        
        if 'STDEV' in formula_text or 'VOLATILITY' in formula_text:
            assessment['risk_metrics_calculated'].append('Volatility Calculation')
        
        # Regulatory compliance indicators
        if assessment['var_modeling'] and assessment['backtesting']:
            assessment['regulatory_compliance'].append('Basel III Market Risk')
        
        if assessment['stress_testing']:
            assessment['regulatory_compliance'].append('CCAR/DFAST Stress Testing')
        
        return assessment
    
    def _identify_model_validation_indicators(self, worksheets: List[WorksheetAnalysis], 
                                            vba_analysis: Dict[str, VBAAnalysis]) -> Dict[str, Any]:
        """Identify model validation and testing indicators"""
        
        indicators = {
            'validation_tests_present': False,
            'sensitivity_analysis': False,
            'benchmark_comparisons': False,
            'data_quality_checks': False,
            'assumption_testing': False,
            'validation_methods': [],
            'testing_coverage': 'basic'
        }
        
        # Check for validation patterns
        all_formulas = []
        for ws in worksheets:
            all_formulas.extend(ws.formulas)
            # Check data validation rules
            if ws.data_validation:
                indicators['data_quality_checks'] = True
                indicators['validation_methods'].append('Data Validation Rules')
        
        formula_text = ' '.join(all_formulas).upper()
        
        # Sensitivity analysis
        if any(term in formula_text for term in ['SENSITIVITY', 'DELTA', 'WHAT.IF']):
            indicators['sensitivity_analysis'] = True
            indicators['validation_methods'].append('Sensitivity Analysis')
        
        # Benchmark comparisons
        if any(term in formula_text for term in ['BENCHMARK', 'COMPARE', 'BASELINE']):
            indicators['benchmark_comparisons'] = True
            indicators['validation_methods'].append('Benchmark Comparison')
        
        # Assumption testing
        if any(term in formula_text for term in ['ASSUMPTION', 'HYPOTHESIS', 'TEST']):
            indicators['assumption_testing'] = True
            indicators['validation_methods'].append('Assumption Testing')
        
        # Determine validation coverage
        method_count = len(indicators['validation_methods'])
        if method_count >= 4:
            indicators['testing_coverage'] = 'comprehensive'
        elif method_count >= 2:
            indicators['testing_coverage'] = 'adequate'
        
        indicators['validation_tests_present'] = method_count > 0
        
        return indicators


# Example usage
if __name__ == "__main__":
    if HAS_OPENPYXL:
        parser = AdvancedExcelParser()
        
        # Test with a sample file
        try:
            analysis = parser.parse("sample.xlsm")
            print(f"Advanced analysis complete: {analysis.filename}")
            print(f"Worksheets: {len(analysis.worksheets)}")
            print(f"Total formulas: {analysis.formulas_summary['total_formulas']}")
            print(f"Financial modeling sophistication: {analysis.financial_modeling_analysis['sophistication_level']}")
            print(f"Model types: {', '.join(analysis.financial_modeling_analysis['model_types_detected'])}")
        except Exception as e:
            print(f"Error: {e}")
    else:
        print("openpyxl not available. Install with: pip install openpyxl")